"""
Report service — downloadable player and team reports.

Reports:
1. All Players List (pre-auction) — name, place, phone, role
2. Post-Auction Player List — all players with status (SOLD/UNSOLD/NOT_PLAYING)
3. Team-wise Player List — each team's squad with sold price + summary
4. All reports can be PDF or Excel (xlsx)
"""
from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from auction.models import Player, Team, TournamentSettings


# ─────────────────────────────────────────────
# Shared PDF helpers
# ─────────────────────────────────────────────

def _doc_styles():
    styles = getSampleStyleSheet()
    title  = ParagraphStyle("T", parent=styles["Title"],
                            fontSize=14, alignment=TA_CENTER, spaceAfter=4,
                            textColor=colors.HexColor("#1a1a2e"))
    sub    = ParagraphStyle("S", parent=styles["Normal"],
                            fontSize=9, alignment=TA_CENTER, spaceAfter=6,
                            textColor=colors.grey)
    h2     = ParagraphStyle("H2", parent=styles["Heading2"],
                            fontSize=11, spaceBefore=8, spaceAfter=3,
                            textColor=colors.HexColor("#1a1a2e"),
                            backColor=colors.HexColor("#ecf0f1"),
                            leftIndent=4)
    normal = styles["Normal"]
    return title, sub, h2, normal


def _base_table_style(header_bg=colors.HexColor("#2c3e50"),
                      header_fg=colors.white,
                      alt_bg=colors.HexColor("#f5f5f5")):
    return [
        ("BACKGROUND",    (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR",     (0, 0), (-1, 0), header_fg),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("LINEBELOW",     (0, 0), (-1, 0), 1.0, header_bg),
    ]


def _alt_rows(style_list, num_rows):
    for i in range(2, num_rows + 1, 2):
        style_list.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f9f9f9")))
    return style_list


# ─────────────────────────────────────────────
# 1. All Players List PDF
# ─────────────────────────────────────────────

def all_players_pdf(role_filter=None):
    """
    role_filter: None = all roles, or "AR"/"BAT"/"BOWL"/"PLY"
    """
    ts     = TournamentSettings.get()
    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               leftMargin=12*mm, rightMargin=12*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    title_s, sub_s, h2_s, normal_s = _doc_styles()
    elements = []

    subtitle = f"Players — {role_filter}" if role_filter else "All Players"
    elements.append(Paragraph(ts.tournament_name, title_s))
    elements.append(Paragraph(subtitle, sub_s))
    elements.append(HRFlowable(width="100%", thickness=0.5,
                               color=colors.HexColor("#cccccc")))
    elements.append(Spacer(1, 4*mm))

    qs = Player.objects.all().order_by("role", "name")
    if role_filter:
        qs = qs.filter(role=role_filter)

    ROLE_LABELS = {"AR": "All Rounder", "BAT": "Batsman",
                   "BOWL": "Bowler", "PLY": "Player"}

    # Group by role
    role_groups = {}
    for p in qs:
        role_groups.setdefault(p.role, []).append(p)

    col_w = [10*mm, 55*mm, 40*mm, 35*mm, 25*mm]

    for role, players in role_groups.items():
        elements.append(Paragraph(f"  {ROLE_LABELS.get(role, role)} ({len(players)})", h2_s))
        data = [["#", "Name", "Place", "Phone", "Base Price"]]
        for p in players:
            data.append([str(p.serial_number), p.name, p.place or "—",
                         p.phone or "—", str(p.base_price)])
        style = _base_table_style()
        _alt_rows(style, len(data) - 1)
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle(style))
        elements.append(t)
        elements.append(Spacer(1, 3*mm))

    elements.append(Paragraph(
        f"Total: {qs.count()} players",
        ParagraphStyle("ft", parent=normal_s, fontSize=8,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────
# 2. Post-Auction Player Status PDF
# ─────────────────────────────────────────────

def auction_players_pdf(status_filter=None):
    """
    status_filter: None=all, "SOLD"/"UNSOLD"/"NOT_PLAYING"/"AVAILABLE"
    """
    ts     = TournamentSettings.get()
    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               leftMargin=12*mm, rightMargin=12*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    title_s, sub_s, h2_s, normal_s = _doc_styles()
    elements = []

    subtitle = f"Auction Results — {status_filter.title()}" if status_filter else "Auction Results — All Players"
    elements.append(Paragraph(ts.tournament_name, title_s))
    elements.append(Paragraph(subtitle, sub_s))
    elements.append(HRFlowable(width="100%", thickness=0.5,
                               color=colors.HexColor("#cccccc")))
    elements.append(Spacer(1, 4*mm))

    qs = Player.objects.all().order_by("status", "role", "name")
    if status_filter:
        qs = qs.filter(status=status_filter)

    STATUS_LABELS = {
        "SOLD":        ("Sold", colors.HexColor("#27ae60")),
        "UNSOLD":      ("Unsold", colors.HexColor("#e67e22")),
        "NOT_PLAYING": ("Not Playing", colors.HexColor("#95a5a6")),
        "AVAILABLE":   ("Available", colors.HexColor("#3498db")),
    }

    col_w = [8*mm, 50*mm, 25*mm, 28*mm, 20*mm, 35*mm, 22*mm]
    data  = [["#", "Name", "Role", "Place", "Phone", "Team", "Sold Price"]]
    for p in qs:
        team_name = p.team.name if p.team else "—"
        price     = str(p.sold_price) if p.sold_price else "—"
        data.append([
            str(p.serial_number), p.name,
            p.role, p.place or "—", p.phone or "—",
            team_name, price
        ])

    style = _base_table_style()
    _alt_rows(style, len(data) - 1)
    # Colour sold price column
    for row_idx, p in enumerate(list(qs), 1):
        if p.status == "SOLD":
            style.append(("TEXTCOLOR", (6, row_idx), (6, row_idx), colors.HexColor("#27ae60")))
        elif p.status == "NOT_PLAYING":
            style.append(("TEXTCOLOR", (1, row_idx), (1, row_idx), colors.grey))

    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle(style))
    elements.append(t)

    # Summary counts
    elements.append(Spacer(1, 4*mm))
    total     = Player.objects.count()
    sold      = Player.objects.filter(status="SOLD").count()
    unsold    = Player.objects.filter(status="UNSOLD").count()
    not_play  = Player.objects.filter(status="NOT_PLAYING").count()
    available = Player.objects.filter(status="AVAILABLE").count()
    elements.append(Paragraph(
        f"Total: {total}  |  Sold: {sold}  |  Unsold: {unsold}  |  Not Playing: {not_play}  |  Available: {available}",
        ParagraphStyle("ft", parent=normal_s, fontSize=8,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────
# 3. Team-wise Squad PDF
# ─────────────────────────────────────────────

def teamwise_pdf():
    ts     = TournamentSettings.get()
    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               leftMargin=12*mm, rightMargin=12*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    title_s, sub_s, h2_s, normal_s = _doc_styles()
    elements = []

    elements.append(Paragraph(ts.tournament_name, title_s))
    elements.append(Paragraph("Team-wise Squad Summary", sub_s))
    elements.append(HRFlowable(width="100%", thickness=0.5,
                               color=colors.HexColor("#cccccc")))
    elements.append(Spacer(1, 4*mm))

    teams = Team.objects.all().order_by("name")
    col_w = [8*mm, 50*mm, 18*mm, 22*mm, 22*mm]

    grand_spent = 0
    grand_count = 0

    for team in teams:
        players = Player.objects.filter(
            team=team, status="SOLD"
        ).order_by("role", "name")
        if not players.exists():
            continue

        spent = sum(p.sold_price or 0 for p in players)
        grand_spent += spent
        grand_count += players.count()

        owner_str = f"  Owner: {team.owners}" if team.owners else ""
        elements.append(Paragraph(
            f"  {team.name}  ·  {team.get_short()}{owner_str}", h2_s
        ))

        # Summary sub-line
        elements.append(Paragraph(
            f"Players: {players.count()}  |  Spent: {spent}  |  Remaining: {team.remaining_points}",
            ParagraphStyle("sm", parent=normal_s, fontSize=8,
                           textColor=colors.HexColor("#666666"), spaceAfter=3)
        ))

        data = [["#", "Player", "Role", "Place", "Sold Price"]]
        for p in players:
            data.append([str(p.serial_number), p.name, p.role, p.place or "—",
                         str(p.sold_price or "—")])

        style = _base_table_style(
            header_bg=colors.HexColor("#34495e"), header_fg=colors.white
        )
        _alt_rows(style, len(data) - 1)
        style.append(("ALIGN", (4, 1), (4, -1), "CENTER"))
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle(style))
        elements.append(t)
        elements.append(Spacer(1, 4*mm))

    # Grand summary
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(
        f"Grand Total — Players Sold: {grand_count}  |  Total Spent: {grand_spent}",
        ParagraphStyle("ft", parent=normal_s, fontSize=9,
                       textColor=colors.HexColor("#2c3e50"), alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────
# Excel (xlsx) versions
# ─────────────────────────────────────────────

def _xl_header_style(ws, row, cols, fill_hex="1a1a2e", font_hex="FFFFFF"):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        fill  = PatternFill("solid", fgColor=fill_hex)
        font  = Font(bold=True, color=font_hex, size=9)
        align = Alignment(horizontal="center", vertical="center")
        thin  = Side(style="thin", color="CCCCCC")
        bdr   = Border(bottom=thin, right=thin)
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.fill  = fill
            cell.font  = font
            cell.alignment = align
            cell.border = bdr
    except Exception:
        pass


def all_players_excel(role_filter=None):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise ImportError("openpyxl required for Excel export. pip install openpyxl")

    ts = TournamentSettings.get()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Players"

    ws.append([ts.tournament_name])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([""])

    headers = ["#", "Name", "Role", "Place", "Phone", "Base Price"]
    ws.append(headers)
    _xl_header_style(ws, 3, range(1, len(headers) + 1))

    qs = Player.objects.all().order_by("role", "name")
    if role_filter:
        qs = qs.filter(role=role_filter)

    for p in qs:
        ws.append([p.serial_number, p.name, p.role, p.place or "", p.phone or "", p.base_price])

    # Column widths
    for col, width in zip("ABCDEF", [5, 30, 10, 22, 16, 12]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def auction_players_excel(status_filter=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required")

    ts = TournamentSettings.get()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auction Results"

    ws.append([ts.tournament_name])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([""])

    headers = ["#", "Name", "Role", "Place", "Phone", "Status", "Team", "Sold Price"]
    ws.append(headers)
    _xl_header_style(ws, 3, range(1, len(headers) + 1))

    qs = Player.objects.all().order_by("status", "role", "name")
    if status_filter:
        qs = qs.filter(status=status_filter)

    STATUS_COLORS = {
        "SOLD":        "27AE60",
        "UNSOLD":      "E67E22",
        "NOT_PLAYING": "95A5A6",
        "AVAILABLE":   "3498DB",
    }

    for p in qs:
        team_name = p.team.name if p.team else ""
        row_data  = [p.serial_number, p.name, p.role, p.place or "", p.phone or "",
                     p.status, team_name, p.sold_price or ""]
        ws.append(row_data)
        # Colour status cell
        color = STATUS_COLORS.get(p.status)
        if color:
            # colour status cell — find its row by position
            ws.cell(row=ws.max_row, column=6).font = Font(
                color=color, bold=True, size=8
            )

    for col, width in zip("ABCDEFGH", [5, 30, 10, 22, 16, 14, 28, 12]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def teamwise_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required")

    ts = TournamentSettings.get()
    wb = openpyxl.Workbook()

    teams = Team.objects.all().order_by("name")
    first = True

    for team in teams:
        players = Player.objects.filter(
            team=team, status="SOLD"
        ).order_by("role", "name")
        if not players.exists():
            continue

        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = team.get_short()[:30]

        ws.append([team.name])
        ws["A1"].font = Font(bold=True, size=13)
        if team.owners:
            ws.append([f"Owner: {team.owners}"])
        spent = sum(p.sold_price or 0 for p in players)
        ws.append([f"Players: {players.count()}  |  Spent: {spent}  |  Remaining: {team.remaining_points}"])
        ws.append([""])

        headers = ["#", "Player", "Role", "Place", "Phone", "Sold Price"]
        ws.append(headers)
        hrow = ws.max_row
        _xl_header_style(ws, hrow, range(1, 7), fill_hex="34495E")

        for p in players:
            ws.append([p.serial_number, p.name, p.role, p.place or "", p.phone or "",
                       p.sold_price or ""])

        for col, width in zip("ABCDEF", [5, 30, 10, 20, 16, 12]):
            ws.column_dimensions[col].width = width

    if first:
        # No teams had players — add empty sheet
        ws = wb.active
        ws.title = "No Data"
        ws.append(["No sold players found"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
