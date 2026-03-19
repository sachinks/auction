"""
jersey_service.py
─────────────────
Jersey size conversion and PDF export.

Size mapping is read dynamically from TournamentConfig.size_mapping (JSON).
Falls back to a safe default map if config is absent.
"""
import json
import logging
from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER

from auction.models import Jersey, ExtraJerseyMember, Team, TournamentSettings

logger = logging.getLogger("system")

# Fallback map — mirrors the TournamentConfig default JSON
_DEFAULT_SIZE_MAP = {
    "36": "XS",
    "38": "S",
    "40": "M",
    "42": "L",
    "44": "XL",
    "46": "XXL",
}


def _load_size_map():
    """
    Load size mapping from TournamentConfig.size_mapping.
    Keys are stored as strings in JSON, so we normalise to int → str.
    Falls back to _DEFAULT_SIZE_MAP if config is absent or invalid.
    """
    try:
        from auction.models import TournamentConfig
        config = TournamentConfig.objects.first()
        if config and config.size_mapping:
            raw = json.loads(config.size_mapping)
            # Normalise: keys may be "38" or 38, values are size labels
            return {str(k): str(v) for k, v in raw.items()}
    except Exception as e:
        logger.warning(f"_load_size_map: could not read config size_mapping — {e}")
    return dict(_DEFAULT_SIZE_MAP)


class JerseyService:

    # ----------------------------------------
    # CONVERT SIZE NUMBER TO TEXT
    # Reads from TournamentConfig each call so admin changes take effect immediately.
    # ----------------------------------------

    def convert_size(self, size_number):
        size_map = _load_size_map()
        return size_map.get(str(size_number), "UNKNOWN")

    # ----------------------------------------
    # CREATE JERSEY RECORD
    # ----------------------------------------

    def create_jersey(self, player, jersey_name, jersey_number, size_number, sponsor):
        size_text = self.convert_size(size_number)
        Jersey.objects.create(
            player=player,
            jersey_name=jersey_name,
            jersey_number=jersey_number,
            size_number=size_number,
            size_text=size_text,
            sponsor=sponsor,
        )

    # ----------------------------------------
    # EXPORT JERSEY LIST PDF (legacy flat list)
    # ----------------------------------------

    def export_pdf(self):
        jerseys = Jersey.objects.select_related("player").all()

        buffer = BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4)

        data = [["Player", "Jersey Name", "Number", "Size", "Sponsor"]]
        for j in jerseys:
            data.append([
                j.player.name,
                j.jersey_name,
                str(j.jersey_number),
                j.size_text,
                j.sponsor or "",
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.black),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))

        doc.build([table])
        buffer.seek(0)
        return buffer

    # ----------------------------------------
    # EXPORT PLAYERS JERSEY PDF — team + category-wise
    # ----------------------------------------

    def _pdf_styles(self):
        styles = getSampleStyleSheet()
        title  = ParagraphStyle("T", parent=styles["Title"],
                                fontSize=13, alignment=TA_CENTER, spaceAfter=2,
                                textColor=colors.HexColor("#1a1a2e"))
        sub    = ParagraphStyle("S", parent=styles["Normal"],
                                fontSize=8, alignment=TA_CENTER, spaceAfter=4,
                                textColor=colors.grey)
        team   = ParagraphStyle("TM", parent=styles["Heading2"],
                                fontSize=11, spaceBefore=8, spaceAfter=2,
                                textColor=colors.white,
                                backColor=colors.HexColor("#2c3e50"),
                                leftIndent=4)
        cat    = ParagraphStyle("CT", parent=styles["Heading3"],
                                fontSize=9, spaceBefore=4, spaceAfter=2,
                                textColor=colors.HexColor("#1a1a2e"),
                                backColor=colors.HexColor("#ecf0f1"),
                                leftIndent=4)
        return title, sub, team, cat

    def _jersey_table(self, rows):
        col_w = [8*mm, 42*mm, 30*mm, 18*mm, 16*mm, 35*mm]
        header = [["#", "Player", "Jersey Name", "No.", "Size", "Sponsor"]]
        data   = header + rows
        style  = [
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i in range(2, len(data), 2):
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f9f9f9")))
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle(style))
        return t

    def export_players_pdf(self):
        ts            = TournamentSettings.get()
        title_s, sub_s, team_s, cat_s = self._pdf_styles()
        ROLE_LABELS   = {"AR": "All Rounders", "BAT": "Batsmen",
                         "BOWL": "Bowlers", "PLY": "Players"}
        ROLE_ORDER    = ["AR", "BAT", "BOWL", "PLY"]

        jersey_map = {}
        for j in Jersey.objects.select_related("player", "player__team").all():
            jersey_map[j.player_id] = j

        teams    = Team.objects.all().order_by("name")
        elements = []
        elements.append(Paragraph(ts.tournament_name, title_s))
        elements.append(Paragraph("Jersey List — Players (Team & Category wise)", sub_s))
        elements.append(HRFlowable(width="100%", thickness=0.5,
                                   color=colors.HexColor("#cccccc")))

        for team in teams:
            from auction.models import Player as _P
            team_jerseys = []
            for role in ROLE_ORDER:
                players = _P.objects.filter(
                    team=team, status=_P.STATUS_SOLD, role=role
                ).order_by("name")
                rows = []
                for p in players:
                    j = jersey_map.get(p.serial_number)
                    rows.append([
                        str(p.serial_number),
                        p.name,
                        j.jersey_name   if j else "—",
                        str(j.jersey_number) if j and j.jersey_number is not None else "—",
                        j.size_text     if j else "—",
                        j.sponsor       if j else "—",
                    ])
                if rows:
                    team_jerseys.append((role, rows))

            if not team_jerseys:
                continue

            elements.append(Spacer(1, 3*mm))
            elements.append(Paragraph(f"  {team.name}", team_s))
            for role, rows in team_jerseys:
                elements.append(Paragraph(
                    f"  {ROLE_LABELS.get(role, role)}  ({len(rows)})", cat_s
                ))
                elements.append(self._jersey_table(rows))
                elements.append(Spacer(1, 2*mm))

        buffer = BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                   leftMargin=12*mm, rightMargin=12*mm,
                                   topMargin=12*mm, bottomMargin=12*mm)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # ----------------------------------------
    # EXPORT ORGANISERS JERSEY PDF
    # ----------------------------------------

    def export_organisers_pdf(self):
        ts            = TournamentSettings.get()
        title_s, sub_s, team_s, cat_s = self._pdf_styles()

        elements = []
        elements.append(Paragraph(ts.tournament_name, title_s))
        elements.append(Paragraph("Jersey List — Organisers & Team Staff", sub_s))
        elements.append(HRFlowable(width="100%", thickness=0.5,
                                   color=colors.HexColor("#cccccc")))

        # ── Team extras grouped by team ──
        from auction.models import Team as _T
        teams = _T.objects.all().order_by("name")
        for team in teams:
            extras = ExtraJerseyMember.objects.filter(
                member_type=ExtraJerseyMember.TYPE_TEAM, team=team
            ).order_by("name")
            if not extras.exists():
                continue
            elements.append(Spacer(1, 3*mm))
            elements.append(Paragraph(f"  {team.name} — Staff", team_s))
            col_w = [8*mm, 40*mm, 30*mm, 30*mm, 18*mm, 16*mm, 35*mm]
            rows  = [["#", "Name", "Role", "Jersey Name", "No.", "Size", "Sponsor"]]
            for i, em in enumerate(extras, 1):
                rows.append([
                    str(i), em.name, em.role_label or "—",
                    em.jersey_name or "—",
                    str(em.jersey_number) if em.jersey_number is not None else "—",
                    em.size_text or "—",
                    em.sponsor or "—",
                ])
            style = [
                ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1), 8),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            for i in range(2, len(rows), 2):
                style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f9f9f9")))
            t = Table(rows, colWidths=col_w, repeatRows=1)
            t.setStyle(TableStyle(style))
            elements.append(t)
            elements.append(Spacer(1, 2*mm))

        # ── Organisers ──
        orgs = ExtraJerseyMember.objects.filter(
            member_type=ExtraJerseyMember.TYPE_ORGANISER
        ).order_by("group_name", "name")
        if orgs.exists():
            elements.append(Spacer(1, 3*mm))
            elements.append(Paragraph("  Organisers", team_s))
            col_w = [8*mm, 40*mm, 40*mm, 30*mm, 18*mm, 16*mm, 35*mm]
            rows  = [["#", "Name", "Group", "Jersey Name", "No.", "Size", "Sponsor"]]
            for i, em in enumerate(orgs, 1):
                rows.append([
                    str(i), em.name, em.group_name or "—",
                    em.jersey_name or "—",
                    str(em.jersey_number) if em.jersey_number is not None else "—",
                    em.size_text or "—",
                    em.sponsor or "—",
                ])
            style = [
                ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, -1), 8),
                ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
                ("TOPPADDING",    (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            for i in range(2, len(rows), 2):
                style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f9f9f9")))
            t = Table(rows, colWidths=col_w, repeatRows=1)
            t.setStyle(TableStyle(style))
            elements.append(t)

        buffer = BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                   leftMargin=12*mm, rightMargin=12*mm,
                                   topMargin=12*mm, bottomMargin=12*mm)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # ----------------------------------------
    # EXPORT PLAYERS JERSEY EXCEL — team + category-wise
    # ----------------------------------------

    def export_players_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl required")

        ts          = TournamentSettings.get()
        ROLE_LABELS = {"AR": "All Rounders", "BAT": "Batsmen",
                       "BOWL": "Bowlers", "PLY": "Players"}
        ROLE_ORDER  = ["AR", "BAT", "BOWL", "PLY"]

        jersey_map = {}
        for j in Jersey.objects.select_related("player").all():
            jersey_map[j.player_id] = j

        wb   = openpyxl.Workbook()
        first = True

        for team in Team.objects.all().order_by("name"):
            from auction.models import Player as _P
            has_any = False
            for role in ROLE_ORDER:
                if _P.objects.filter(team=team, status=_P.STATUS_SOLD, role=role).exists():
                    has_any = True
                    break
            if not has_any:
                continue

            ws = wb.active if first else wb.create_sheet()
            first = False
            ws.title = team.get_short()[:30]

            ws.append([ts.tournament_name])
            ws["A1"].font = Font(bold=True, size=13)
            ws.append([team.name])
            ws["A2"].font = Font(bold=True, size=11)
            ws.append([""])

            header_fill = PatternFill("solid", fgColor="2C3E50")
            header_font = Font(bold=True, color="FFFFFF", size=9)

            row_num = 4
            for role in ROLE_ORDER:
                players = _P.objects.filter(
                    team=team, status=_P.STATUS_SOLD, role=role
                ).order_by("name")
                if not players.exists():
                    continue

                # Category heading
                ws.append([ROLE_LABELS.get(role, role)])
                cat_cell = ws.cell(row=row_num, column=1)
                cat_cell.font = Font(bold=True, size=10, color="1A1A2E")
                cat_cell.fill = PatternFill("solid", fgColor="ECF0F1")
                row_num += 1

                # Header row
                headers = ["#", "Player", "Jersey Name", "Jersey No.", "Size No.", "Size", "Sponsor"]
                ws.append(headers)
                for col in range(1, len(headers) + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center")
                row_num += 1

                for i, p in enumerate(players, 1):
                    j = jersey_map.get(p.serial_number)
                    ws.append([
                        i, p.name,
                        j.jersey_name   if j else "",
                        j.jersey_number if j and j.jersey_number is not None else "",
                        j.size_number   if j and j.size_number   is not None else "",
                        j.size_text     if j else "",
                        j.sponsor       if j else "",
                    ])
                    if row_num % 2 == 0:
                        for col in range(1, 8):
                            ws.cell(row=row_num, column=col).fill = PatternFill(
                                "solid", fgColor="F9F9F9"
                            )
                    row_num += 1

                ws.append([""])
                row_num += 1

            for col, width in zip("ABCDEFG", [5, 28, 22, 12, 10, 10, 28]):
                ws.column_dimensions[chr(64 + ord('A') - ord('A') + col)].width = width
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 10
            ws.column_dimensions["G"].width = 28

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ----------------------------------------
    # EXPORT ORGANISERS JERSEY EXCEL
    # ----------------------------------------

    def export_organisers_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl required")

        ts = TournamentSettings.get()
        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(bold=True, color="FFFFFF", size=9)

        # ── Team extras — one sheet per team ──
        first = True
        for team in Team.objects.all().order_by("name"):
            extras = ExtraJerseyMember.objects.filter(
                member_type=ExtraJerseyMember.TYPE_TEAM, team=team
            ).order_by("name")
            if not extras.exists():
                continue

            ws = wb.active if first else wb.create_sheet()
            first = False
            ws.title = (team.get_short() + " Staff")[:30]

            ws.append([ts.tournament_name])
            ws["A1"].font = Font(bold=True, size=13)
            ws.append([f"{team.name} — Staff"])
            ws["A2"].font = Font(bold=True, size=11)
            ws.append([""])

            headers = ["#", "Name", "Role", "Jersey Name", "Jersey No.", "Size No.", "Size", "Sponsor"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=4, column=col)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")

            for i, em in enumerate(extras, 1):
                ws.append([
                    i, em.name, em.role_label or "",
                    em.jersey_name or "",
                    em.jersey_number if em.jersey_number is not None else "",
                    em.size_number   if em.size_number   is not None else "",
                    em.size_text or "", em.sponsor or "",
                ])

            for col, width in zip("ABCDEFGH", [5, 28, 18, 22, 12, 10, 10, 28]):
                ws.column_dimensions[col].width = width

        # ── Organisers sheet ──
        orgs = ExtraJerseyMember.objects.filter(
            member_type=ExtraJerseyMember.TYPE_ORGANISER
        ).order_by("group_name", "name")
        if orgs.exists():
            ws = wb.active if first else wb.create_sheet(title="Organisers")
            if first:
                ws.title = "Organisers"
            first = False

            ws.append([ts.tournament_name])
            ws["A1"].font = Font(bold=True, size=13)
            ws.append(["Organisers"])
            ws["A2"].font = Font(bold=True, size=11)
            ws.append([""])

            headers = ["#", "Name", "Group", "Jersey Name", "Jersey No.", "Size No.", "Size", "Sponsor"]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=4, column=col)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")

            for i, em in enumerate(orgs, 1):
                ws.append([
                    i, em.name, em.group_name or "",
                    em.jersey_name or "",
                    em.jersey_number if em.jersey_number is not None else "",
                    em.size_number   if em.size_number   is not None else "",
                    em.size_text or "", em.sponsor or "",
                ])

            for col, width in zip("ABCDEFGH", [5, 28, 28, 22, 12, 10, 10, 28]):
                ws.column_dimensions[col].width = width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
